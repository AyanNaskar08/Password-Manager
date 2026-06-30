"""
Password Manager — Flask Web Application
"""

import os
import csv
import io
import random
import string
from datetime import datetime, timezone

# pyrefly: ignore [missing-import]
import bcrypt
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, flash, send_file
)

from models import db, MasterPassword, PasswordEntry
from crypto_utils import encrypt_password, decrypt_password

app = Flask(__name__)
app.secret_key = os.urandom(32)

# Database configuration
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'vault.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

db.init_app(app)

with app.app_context():
    db.create_all()


# ── Auth helpers ────────────────────────────────────────────────

def is_setup_done():
    """Check if the master password has been set up."""
    return MasterPassword.query.first() is not None


def is_logged_in():
    """Check if the user is currently authenticated."""
    return session.get('authenticated', False)


def login_required(f):
    """Decorator to protect routes behind authentication."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not is_setup_done():
            return redirect(url_for('setup'))
        if not is_logged_in():
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ── Page Routes ─────────────────────────────────────────────────

@app.route('/')
def index():
    if not is_setup_done():
        return redirect(url_for('setup'))
    if is_logged_in():
        return redirect(url_for('vault'))
    return redirect(url_for('login'))


@app.route('/setup', methods=['GET', 'POST'])
def setup():
    if is_setup_done():
        return redirect(url_for('login'))

    if request.method == 'POST':
        master_pwd = request.form.get('master_password', '')
        confirm_pwd = request.form.get('confirm_password', '')

        if len(master_pwd) < 6:
            flash('Master password must be at least 6 characters.', 'error')
            return render_template('setup.html')

        if master_pwd != confirm_pwd:
            flash('Passwords do not match.', 'error')
            return render_template('setup.html')

        hashed = bcrypt.hashpw(master_pwd.encode(), bcrypt.gensalt())
        master = MasterPassword(password_hash=hashed.decode())
        db.session.add(master)
        db.session.commit()

        flash('Master password created! Please log in.', 'success')
        return redirect(url_for('login'))

    return render_template('setup.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if not is_setup_done():
        return redirect(url_for('setup'))
    if is_logged_in():
        return redirect(url_for('vault'))

    if request.method == 'POST':
        master_pwd = request.form.get('master_password', '')
        master = MasterPassword.query.first()

        if bcrypt.checkpw(master_pwd.encode(), master.password_hash.encode()):
            session['authenticated'] = True
            session['master_key'] = master_pwd  # kept in session for encryption
            return redirect(url_for('vault'))
        else:
            flash('Incorrect master password.', 'error')

    return render_template('login.html')


@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/vault')
@login_required
def vault():
    return render_template('vault.html')


# ── API Routes ──────────────────────────────────────────────────

@app.route('/api/passwords', methods=['GET'])
@login_required
def get_passwords():
    search = request.args.get('search', '').strip().lower()
    category = request.args.get('category', '').strip()
    master_key = session.get('master_key', '')

    query = PasswordEntry.query

    if search:
        query = query.filter(
            db.or_(
                PasswordEntry.website.ilike(f'%{search}%'),
                PasswordEntry.username.ilike(f'%{search}%'),
                PasswordEntry.notes.ilike(f'%{search}%'),
            )
        )

    if category and category != 'All':
        query = query.filter(PasswordEntry.category == category)

    entries = query.order_by(PasswordEntry.created_at.desc()).all()

    results = []
    for entry in entries:
        try:
            decrypted = decrypt_password(entry.encrypted_password, master_key)
        except Exception:
            decrypted = '[decryption error]'
        results.append(entry.to_dict(decrypted_password=decrypted))

    return jsonify(results)


@app.route('/api/passwords', methods=['POST'])
@login_required
def add_password():
    data = request.get_json()
    master_key = session.get('master_key', '')

    website = data.get('website', '').strip()
    username = data.get('username', '').strip()
    password_plain = data.get('password', '').strip()
    category = data.get('category', 'General').strip()
    notes = data.get('notes', '').strip()

    if not website or not password_plain:
        return jsonify({'error': 'Website and password are required.'}), 400

    encrypted = encrypt_password(password_plain, master_key)

    entry = PasswordEntry(
        website=website,
        username=username,
        encrypted_password=encrypted,
        category=category,
        notes=notes,
    )
    db.session.add(entry)
    db.session.commit()

    try:
        decrypted = decrypt_password(entry.encrypted_password, master_key)
    except Exception:
        decrypted = password_plain

    return jsonify(entry.to_dict(decrypted_password=decrypted)), 201


@app.route('/api/passwords/<int:entry_id>', methods=['PUT'])
@login_required
def update_password(entry_id):
    data = request.get_json()
    master_key = session.get('master_key', '')

    entry = PasswordEntry.query.get_or_404(entry_id)

    if 'website' in data:
        entry.website = data['website'].strip()
    if 'username' in data:
        entry.username = data['username'].strip()
    if 'password' in data and data['password'].strip():
        entry.encrypted_password = encrypt_password(data['password'].strip(), master_key)
    if 'category' in data:
        entry.category = data['category'].strip()
    if 'notes' in data:
        entry.notes = data['notes'].strip()

    entry.updated_at = datetime.now(timezone.utc)
    db.session.commit()

    try:
        decrypted = decrypt_password(entry.encrypted_password, master_key)
    except Exception:
        decrypted = '[decryption error]'

    return jsonify(entry.to_dict(decrypted_password=decrypted))


@app.route('/api/passwords/<int:entry_id>', methods=['DELETE'])
@login_required
def delete_password(entry_id):
    entry = PasswordEntry.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    return jsonify({'message': 'Deleted successfully.'})


@app.route('/api/generate', methods=['POST'])
@login_required
def generate_password():
    data = request.get_json() or {}
    length = min(max(int(data.get('length', 16)), 4), 128)
    use_upper = data.get('uppercase', True)
    use_lower = data.get('lowercase', True)
    use_digits = data.get('digits', True)
    use_symbols = data.get('symbols', True)

    chars = ''
    if use_upper:
        chars += string.ascii_uppercase
    if use_lower:
        chars += string.ascii_lowercase
    if use_digits:
        chars += string.digits
    if use_symbols:
        chars += '!@#$%^&*()-_=+[]{}|;:,.<>?'

    if not chars:
        chars = string.ascii_letters + string.digits

    pwd = ''.join(random.SystemRandom().choice(chars) for _ in range(length))
    return jsonify({'password': pwd})


@app.route('/api/import', methods=['POST'])
@login_required
def import_file():
    master_key = session.get('master_key', '')

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided.'}), 400

    file = request.files['file']
    filename = file.filename.lower()

    entries = []

    try:
        if filename.endswith('.csv'):
            stream = io.StringIO(file.stream.read().decode('utf-8'))
            reader = csv.DictReader(stream)
            for row in reader:
                # Flexible column name mapping
                website = (row.get('website') or row.get('url') or row.get('site')
                           or row.get('name') or row.get('Website') or row.get('URL')
                           or row.get('Name') or row.get('Site') or '').strip()
                username = (row.get('username') or row.get('user') or row.get('email')
                            or row.get('login') or row.get('Username') or row.get('User')
                            or row.get('Email') or row.get('Login') or '').strip()
                password_val = (row.get('password') or row.get('pass') or row.get('pwd')
                                or row.get('Password') or row.get('Pass') or '').strip()
                category = (row.get('category') or row.get('group') or row.get('folder')
                            or row.get('Category') or row.get('Group') or 'Imported').strip()
                notes = (row.get('notes') or row.get('note') or row.get('Notes') or '').strip()

                if website and password_val:
                    entries.append({
                        'website': website,
                        'username': username,
                        'password': password_val,
                        'category': category or 'Imported',
                        'notes': notes,
                    })

        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            from openpyxl import load_workbook
            wb = load_workbook(file, read_only=True)
            ws = wb.active

            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip().lower() if h else '' for h in row]
                    continue

                row_dict = {}
                for j, val in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = str(val).strip() if val else ''

                website = (row_dict.get('website') or row_dict.get('url')
                           or row_dict.get('site') or row_dict.get('name') or '').strip()
                username = (row_dict.get('username') or row_dict.get('user')
                            or row_dict.get('email') or row_dict.get('login') or '').strip()
                password_val = (row_dict.get('password') or row_dict.get('pass')
                                or row_dict.get('pwd') or '').strip()
                category = (row_dict.get('category') or row_dict.get('group')
                            or row_dict.get('folder') or 'Imported').strip()
                notes = (row_dict.get('notes') or row_dict.get('note') or '').strip()

                if website and password_val:
                    entries.append({
                        'website': website,
                        'username': username,
                        'password': password_val,
                        'category': category or 'Imported',
                        'notes': notes,
                    })
            wb.close()
        else:
            return jsonify({'error': 'Unsupported file format. Use .csv or .xlsx'}), 400

    except Exception as e:
        return jsonify({'error': f'Error parsing file: {str(e)}'}), 400

    # Save all entries
    count = 0
    for entry_data in entries:
        encrypted = encrypt_password(entry_data['password'], master_key)
        entry = PasswordEntry(
            website=entry_data['website'],
            username=entry_data['username'],
            encrypted_password=encrypted,
            category=entry_data['category'],
            notes=entry_data['notes'],
        )
        db.session.add(entry)
        count += 1

    db.session.commit()
    return jsonify({'message': f'Successfully imported {count} passwords.', 'count': count})


@app.route('/api/export', methods=['GET'])
@login_required
def export_passwords():
    master_key = session.get('master_key', '')
    entries = PasswordEntry.query.order_by(PasswordEntry.created_at.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Website', 'Username', 'Password', 'Category', 'Notes'])

    for entry in entries:
        try:
            decrypted = decrypt_password(entry.encrypted_password, master_key)
        except Exception:
            decrypted = ''
        writer.writerow([entry.website, entry.username, decrypted, entry.category, entry.notes])

    output.seek(0)
    mem = io.BytesIO(output.getvalue().encode('utf-8'))
    mem.seek(0)

    return send_file(
        mem,
        mimetype='text/csv',
        as_attachment=True,
        download_name='passwords_export.csv',
    )


@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    results = db.session.query(PasswordEntry.category).distinct().all()
    categories = sorted(set(r[0] for r in results if r[0]))
    return jsonify(categories)


if __name__ == '__main__':
    app.run(debug=True, port=5000)
